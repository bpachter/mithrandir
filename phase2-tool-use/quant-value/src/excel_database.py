"""
Excel Database Manager for Quantitative Value Model

This module manages the single Excel workbook that serves as the primary database
for the quantitative value investment system.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
import logging
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ExcelDatabase:
    """Manages the quantitative value Excel database workbook."""

    def __init__(self, database_path: Path = None):
        """Initialize Excel database manager.

        Args:
            database_path: Path to the Excel database file
        """
        if database_path is None:
            database_path = Path('data/reports/quantitative_value_database.xlsx')

        self.database_path = database_path
        self.database_path.parent.mkdir(parents=True, exist_ok=True)

    def create_database(self,
                       ttm_df: pd.DataFrame,
                       annual_df: Optional[pd.DataFrame] = None,
                       quarterly_df: Optional[pd.DataFrame] = None,
                       portfolio_df: Optional[pd.DataFrame] = None,
                       model_criteria: Optional[Dict] = None,
                       franchise_power_df: Optional[pd.DataFrame] = None) -> None:
        """Create or update the Excel database with all sheets.

        Args:
            ttm_df: DataFrame with TTM metrics (primary screening database)
            annual_df: DataFrame with annual metrics (optional, supplementary)
            quarterly_df: DataFrame with quarterly metrics (optional, supplementary)
            portfolio_df: DataFrame with portfolio holdings (optional)
            model_criteria: Dictionary of model criteria rules (optional)
            franchise_power_df: DataFrame with Franchise Power metrics (optional)
        """
        logger.info(f"Creating Excel database at {self.database_path}")

        with pd.ExcelWriter(self.database_path, engine='openpyxl') as writer:
            # Sheet 1: Universe_TTM (TTM metrics - PRIMARY for screening)
            self._write_universe_sheet(writer, ttm_df, sheet_name='Universe_TTM')

            # Sheet 2: Universe_Annual (supplementary reference data)
            if annual_df is not None and not annual_df.empty:
                self._write_universe_sheet(writer, annual_df, sheet_name='Universe_Annual')

            # Sheet 3: Universe_Quarterly (supplementary reference data)
            if quarterly_df is not None and not quarterly_df.empty:
                # Limit to most recent 8 quarters per company for readability
                recent_quarterly = quarterly_df.sort_values('period_end').groupby('ticker').tail(8)
                self._write_universe_sheet(writer, recent_quarterly, sheet_name='Universe_Quarterly')

            # Sheet 4: Franchise Power (8-year quality metrics)
            if franchise_power_df is not None and not franchise_power_df.empty:
                self._write_franchise_power_sheet(writer, franchise_power_df)

            # Sheet 5: Portfolio (user's holdings)
            self._write_portfolio_sheet(writer, portfolio_df)

            # Sheet 6: Portfolio Monitor (check holdings against model)
            if portfolio_df is not None and not portfolio_df.empty:
                self._write_portfolio_monitor_sheet(writer, ttm_df, portfolio_df)

            # Sheet 7: Model Criteria (the rules)
            self._write_model_criteria_sheet(writer, model_criteria)

            # Sheet 8: Metadata (refresh info, stats)
            self._write_metadata_sheet(writer, ttm_df, annual_df, quarterly_df, franchise_power_df)

        logger.info(f"Excel database created successfully with {len(ttm_df)} companies (TTM)")

    def _write_universe_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame,
                             sheet_name: str = 'Universe') -> None:
        """Write a Universe sheet with all companies and metrics."""
        logger.info(f"Writing {sheet_name} sheet with {len(df)} companies")

        # Sort by ticker for easy lookup
        df = df.sort_values('ticker', na_position='last')

        # Write to Excel
        df.to_excel(writer, sheet_name=sheet_name, index=False, freeze_panes=(1, 0))

        # Format the sheet
        worksheet = writer.sheets[sheet_name]
        self._format_header(worksheet)
        self._autofit_columns(worksheet, df)

    def _write_franchise_power_sheet(self, writer: pd.ExcelWriter, franchise_power_df: pd.DataFrame) -> None:
        """Write Franchise Power sheet with 8-year quality metrics."""
        logger.info(f"Writing Franchise_Power sheet with {len(franchise_power_df)} companies")

        # Sort by franchise power score (descending - best first)
        fp_df = franchise_power_df.sort_values('p_franchise_power', ascending=False, na_position='last')

        # Select key columns for display
        display_columns = [
            'ticker', 'cik', 'years_available',
            '8yr_roa', '8yr_roc', 'fcfa', 'margin_growth', 'margin_stability',
            'p_8yr_roa', 'p_8yr_roc', 'p_fcfa', 'p_margin_growth', 'p_margin_stability',
            'p_margin_max', 'fp_score_raw', 'p_franchise_power'
        ]

        # Only include columns that exist
        available_columns = [col for col in display_columns if col in fp_df.columns]
        fp_export = fp_df[available_columns].copy()

        # Write to Excel
        fp_export.to_excel(writer, sheet_name='Franchise_Power', index=False, freeze_panes=(1, 0))

        # Format the sheet
        worksheet = writer.sheets['Franchise_Power']
        self._format_header(worksheet)
        self._autofit_columns(worksheet, fp_export)

        # Add descriptive note at the top (as a separate row above data)
        # This will require shifting the data down, which we'll skip for now
        # Just log the explanation
        logger.info("Franchise Power metrics capture 8-year historical quality:")
        logger.info("  - 8yr_ROA: Geometric average Return on Assets")
        logger.info("  - 8yr_ROC: Geometric average Return on Capital")
        logger.info("  - FCFA: Free Cash Flow to Assets (cumulative)")
        logger.info("  - Margin Growth: 8-year CAGR of gross margins")
        logger.info("  - Margin Stability: Mean margin / Std deviation")
        logger.info("  - p_franchise_power: Final percentile score (0-100, higher = better)")

    def _write_portfolio_sheet(self, writer: pd.ExcelWriter, portfolio_df: Optional[pd.DataFrame]) -> None:
        """Write the Portfolio sheet with user's holdings."""
        if portfolio_df is None or portfolio_df.empty:
            # Create template portfolio sheet
            template_df = pd.DataFrame({
                'ticker': ['AAPL', 'GOOGL', 'MSFT'],  # Examples
                'shares': [100, 50, 75],
                'cost_basis': [150.00, 2800.00, 380.00],
                'purchase_date': ['2024-01-15', '2024-02-20', '2024-03-10'],
                'notes': ['Example holding', 'Example holding', 'Example holding']
            })
            template_df.to_excel(writer, sheet_name='Portfolio', index=False)
            logger.info("Created template Portfolio sheet")
        else:
            portfolio_df.to_excel(writer, sheet_name='Portfolio', index=False)
            logger.info(f"Writing Portfolio sheet with {len(portfolio_df)} holdings")

        worksheet = writer.sheets['Portfolio']
        self._format_header(worksheet)

    def _write_portfolio_monitor_sheet(self, writer: pd.ExcelWriter,
                                      universe_df: pd.DataFrame,
                                      portfolio_df: pd.DataFrame) -> None:
        """Write Portfolio Monitor sheet - checks if holdings still meet criteria."""
        logger.info("Writing Portfolio Monitor sheet")

        # Merge portfolio with universe data to get current metrics
        monitor_df = portfolio_df.merge(
            universe_df,
            on='ticker',
            how='left',
            suffixes=('_portfolio', '_current')
        )

        # Add monitoring columns (will be populated based on model criteria later)
        monitor_df['meets_criteria'] = 'TBD'  # To be determined after model is defined
        monitor_df['last_check_date'] = datetime.now().strftime('%Y-%m-%d')

        # Select relevant columns for monitoring
        monitor_columns = ['ticker', 'shares', 'cost_basis', 'purchase_date',
                          'meets_criteria', 'last_check_date', 'notes']

        # Add key metrics if they exist
        metric_columns = ['f_score', 'value_composite', 'net_margin', 'roe', 'roa',
                         'revenue_yoy', 'fcf', 'current_ratio']
        for col in metric_columns:
            if col in monitor_df.columns:
                monitor_columns.append(col)

        # Write to Excel
        monitor_df[monitor_columns].to_excel(writer, sheet_name='Portfolio_Monitor', index=False)

        worksheet = writer.sheets['Portfolio_Monitor']
        self._format_header(worksheet)

    def _write_model_criteria_sheet(self, writer: pd.ExcelWriter,
                                   model_criteria: Optional[Dict]) -> None:
        """Write Model Criteria sheet - the rules for stock selection."""
        logger.info("Writing Model Criteria sheet")

        if model_criteria is None:
            # Create template criteria sheet
            criteria_data = {
                'Criterion': [
                    '=== TO BE DEFINED ===',
                    'Minimum F-Score',
                    'Maximum Value Composite Rank',
                    'Minimum Market Cap',
                    'Minimum ROE',
                    'Minimum Current Ratio',
                    'Revenue Growth (YoY)',
                    'Free Cash Flow Positive',
                    'Debt to Equity Max',
                    'Other criteria...'
                ],
                'Threshold': [
                    '',
                    '7',
                    '50',
                    '$100M',
                    '10%',
                    '1.5',
                    '>0%',
                    'YES',
                    '<2.0',
                    'TBD'
                ],
                'Notes': [
                    'Define your Quantitative Value model rules here',
                    'Piotroski F-Score (0-9)',
                    'Lower is better (cheaper stocks)',
                    'Exclude micro-caps',
                    'Return on Equity threshold',
                    'Liquidity requirement',
                    'Positive revenue growth required',
                    'Must generate positive FCF',
                    'Maximum leverage ratio',
                    'Add additional criteria as needed'
                ]
            }
        else:
            criteria_data = {
                'Criterion': list(model_criteria.keys()),
                'Threshold': list(model_criteria.values()),
                'Notes': [''] * len(model_criteria)
            }

        criteria_df = pd.DataFrame(criteria_data)
        criteria_df.to_excel(writer, sheet_name='Model_Criteria', index=False)

        worksheet = writer.sheets['Model_Criteria']
        self._format_header(worksheet)

        # Highlight the header row in yellow
        for cell in worksheet[1]:
            cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    def _write_metadata_sheet(self, writer: pd.ExcelWriter,
                             ttm_df: pd.DataFrame,
                             annual_df: Optional[pd.DataFrame] = None,
                             quarterly_df: Optional[pd.DataFrame] = None,
                             franchise_power_df: Optional[pd.DataFrame] = None) -> None:
        """Write Metadata sheet with refresh info and statistics."""
        logger.info("Writing Metadata sheet")

        # Calculate statistics
        stats = {
            'Metric': [
                'Last Data Refresh',
                'Data Type',
                'Total Companies (TTM)',
                'Total Companies (Annual)',
                'Total Companies (Quarterly)',
                'Total Companies (Franchise Power)',
                'Companies with Complete TTM Data',
                'Companies with Franchise Power Score',
                'Average F-Score (TTM)',
                'Median F-Score (TTM)',
                'Companies with F-Score >= 7',
                'Average Franchise Power Score',
                'Median Franchise Power Score',
                'Companies with FP >= 75th percentile',
                'Annual Data: Years of History',
                'Quarterly Data: Years of History',
                'Franchise Power: Years of History',
                'Data Source',
                'Database Version'
            ],
            'Value': [
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'Trailing Twelve Months (TTM)',
                len(ttm_df),
                len(annual_df) if annual_df is not None else 'N/A',
                len(quarterly_df) if quarterly_df is not None else 'N/A',
                len(franchise_power_df) if franchise_power_df is not None else 'N/A',
                ttm_df.dropna(subset=['ticker']).shape[0],
                franchise_power_df['p_franchise_power'].notna().sum() if franchise_power_df is not None and 'p_franchise_power' in franchise_power_df.columns else 'N/A',
                f"{ttm_df['f_score'].mean():.2f}" if 'f_score' in ttm_df.columns else 'N/A',
                f"{ttm_df['f_score'].median():.2f}" if 'f_score' in ttm_df.columns else 'N/A',
                ttm_df[ttm_df['f_score'] >= 7].shape[0] if 'f_score' in ttm_df.columns else 'N/A',
                f"{franchise_power_df['p_franchise_power'].mean():.2f}" if franchise_power_df is not None and 'p_franchise_power' in franchise_power_df.columns else 'N/A',
                f"{franchise_power_df['p_franchise_power'].median():.2f}" if franchise_power_df is not None and 'p_franchise_power' in franchise_power_df.columns else 'N/A',
                franchise_power_df[franchise_power_df['p_franchise_power'] >= 75].shape[0] if franchise_power_df is not None and 'p_franchise_power' in franchise_power_df.columns else 'N/A',
                '2 years',
                '8 years (32 quarters)',
                '8 years (annual)',
                'SEC EDGAR (companyfacts API)',
                '3.0 (TTM + Franchise Power)'
            ]
        }

        metadata_df = pd.DataFrame(stats)
        metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

        worksheet = writer.sheets['Metadata']
        self._format_header(worksheet)

    def _format_header(self, worksheet) -> None:
        """Format the header row with bold font and background color."""
        for cell in worksheet[1]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def _autofit_columns(self, worksheet, df: pd.DataFrame, max_width: int = 50) -> None:
        """Auto-fit column widths based on content."""
        for idx, col in enumerate(df.columns, 1):
            # Get max length of column content
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(str(col))
            )
            # Add some padding and cap at max_width
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = adjusted_width

    def load_portfolio(self) -> pd.DataFrame:
        """Load portfolio from the Excel database.

        Returns:
            DataFrame with portfolio holdings
        """
        if not self.database_path.exists():
            logger.warning(f"Database not found at {self.database_path}")
            return pd.DataFrame()

        try:
            portfolio_df = pd.read_excel(self.database_path, sheet_name='Portfolio')
            logger.info(f"Loaded {len(portfolio_df)} portfolio holdings")
            return portfolio_df
        except Exception as e:
            logger.error(f"Error loading portfolio: {e}")
            return pd.DataFrame()

    def load_universe(self, sheet_name: str = 'Universe_TTM') -> pd.DataFrame:
        """Load universe data from the Excel database.

        Args:
            sheet_name: Name of the sheet to load (default: Universe_TTM)

        Returns:
            DataFrame with all companies and metrics
        """
        if not self.database_path.exists():
            logger.warning(f"Database not found at {self.database_path}")
            return pd.DataFrame()

        try:
            universe_df = pd.read_excel(self.database_path, sheet_name=sheet_name)
            logger.info(f"Loaded {len(universe_df)} companies from {sheet_name}")
            return universe_df
        except Exception as e:
            logger.error(f"Error loading {sheet_name}: {e}")
            return pd.DataFrame()


def create_excel_database_from_csvs():
    """Create the Excel database from existing CSV files with TTM calculations."""
    from config import get_config
    from ttm_calculator import TTMCalculator

    config = get_config()

    # Load the processed data
    metrics_path = config.get_metrics_path()

    if not metrics_path.exists():
        logger.error(f"Metrics file not found at {metrics_path}")
        logger.info("Please run the data pipeline first: python src/run_all.py")
        return

    # Load metrics (includes all fundamentals + derived metrics)
    logger.info(f"Loading metrics from {metrics_path}")
    metrics_df = pd.read_csv(metrics_path)

    # Calculate TTM metrics
    logger.info("Calculating TTM metrics...")
    calculator = TTMCalculator()
    ttm_df = calculator.calculate_ttm(metrics_df)

    if ttm_df.empty:
        logger.error("Failed to calculate TTM metrics")
        return

    # Separate annual and quarterly data for supplementary sheets
    annual_df = metrics_df[metrics_df['frequency'] == 'annual'].copy()
    latest_annual = annual_df.sort_values('period_end').groupby('ticker').tail(1)

    quarterly_df = metrics_df[metrics_df['frequency'] == 'quarterly'].copy()

    logger.info(f"TTM: {len(ttm_df)} companies")
    logger.info(f"Annual: {len(latest_annual)} companies")
    logger.info(f"Quarterly: {len(quarterly_df)} records")

    # Create the Excel database
    db = ExcelDatabase()
    db.create_database(
        ttm_df=ttm_df,
        annual_df=latest_annual,
        quarterly_df=quarterly_df,
        portfolio_df=None,  # Will create template
        model_criteria=None  # Will create template
    )

    logger.info(f"Excel database created at: {db.database_path}")


if __name__ == '__main__':
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    create_excel_database_from_csvs()
